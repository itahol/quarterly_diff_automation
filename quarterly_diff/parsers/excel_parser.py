from __future__ import annotations

import os.path
from typing import TYPE_CHECKING, Dict, Tuple
from functools import cached_property

import openpyxl
import xlrd

from quarterly_diff.parsers.company_investment import CompanyInvestment

if TYPE_CHECKING:
    from typing import Generator, Callable, Any, List
    from openpyxl.worksheet.worksheet import Worksheet as XLSXWorksheet
    from openpyxl.worksheet.worksheet import Worksheet as XLSXWorksheet
    from openpyxl.cell import Cell as XLSXCell
    from xlrd.sheet import Sheet as XLSWorksheet

InvestmentPortfolio = Dict[Tuple[str, str], CompanyInvestment]


class ExcelParser:
    @cached_property
    def headers_row_idx(self):
        row_index = self._find_value_row_index("שם המנפיק/שם נייר ערך", "שם המנפיק / שם נייר ערך", 'שם נ"ע')
        return row_index if self._file_ext == ".xls" else row_index + 1

    @cached_property
    def headers_row(self):
        return self._sheet[self.headers_row_idx]

    @cached_property
    def company_name_idx(self):
        return self._find_value_index(self.headers_row, "שם המנפיק/שם נייר ערך", "שם המנפיק / שם נייר ערך", 'שם נ"ע')

    @cached_property
    def company_id_idx(self):
        return self._find_value_index(self.headers_row, "מספר מנפיק")

    @cached_property
    def company_category_idx(self):
        return self._find_value_index(self.headers_row, "ענף מסחר")

    @cached_property
    def nominal_value_idx(self):
        return self._find_value_index(self.headers_row, "ערך נקוב")

    @cached_property
    def currency_col(self):
        return self._find_value_index(self.headers_row, "סוג מטבע")

    @cached_property
    def fair_value_idx(self):
        return self._find_value_index(self.headers_row, "שווי הוגן", "שווי שוק")

    EXT_TO_LIB = {
        ".xlsx": openpyxl,
        ".xls": xlrd,
    }
    COMPANIES_START_ROW_IDX = 12
    STAKE_SHEET_NAME = "לא סחיר - מניות"

    def __init__(self, workbook_path):
        self._file_ext = os.path.splitext(workbook_path)[-1]
        if self._file_ext == ".xls":
            self._workbook = xlrd.open_workbook(workbook_path)
            self._sheet: XLSWorksheet = self._workbook.sheet_by_name(self.STAKE_SHEET_NAME)
        elif self._file_ext == ".xlsx":
            self._workbook = openpyxl.load_workbook(workbook_path)
            self._sheet: XLSXWorksheet = self._workbook[self.STAKE_SHEET_NAME]  # type: ignore[no-redef]
        else:
            raise ValueError(f"Only .xls and .xlsx files are supported - a {self._file_ext} file was provided")

    @staticmethod
    def _find_value_index(row, *values):
        for i, cell in enumerate(row):
            if cell.value and any(cell.value.startswith(value) for value in values):
                return i
        raise ValueError(f"None of the values {values} were found in row")

    def _find_value_row_index(self, *values):
        if self._file_ext == ".xls":
            rows_iter = (self._sheet.row_values(index) for index in range(self._sheet.nrows))
        else:
            rows_iter = self._sheet.iter_rows(values_only=True)

        for index, row in enumerate(rows_iter):
            if any(value in row for value in values):
                return index
        raise ValueError(f"None of the values {values} were found in sheet")

    @staticmethod
    def _get_rows_from_xls(
            sheet: XLSWorksheet,
            start_index: int,
            end_index: int,
            condition_func: Callable[[list], Any]) -> Generator[list, None, None]:
        return (sheet.row(index) for index in range(start_index, end_index) if condition_func(sheet.row(index)))

    @staticmethod
    def _get_rows_from_xlsx(sheet: XLSXWorksheet,
                            start_index: int,
                            end_index: int,
                            condition_func: Callable[[list], Any]) -> Generator[List[XLSXCell], None, None]:
        return (list(row) for row in sheet.iter_rows(min_row=start_index, max_row=end_index) if
                condition_func(list(row)))

    def _get_company_id(self, investment: list) -> str:
        cell_value = investment[self.company_id_idx].value
        if isinstance(cell_value, str):
            if "תא ללא תוכן" in cell_value:
                cell_value = ""
            else:
                cell_value = cell_value.strip()
        return cell_value

    def _get_nominal_value(self, investment: list) -> float:
        return investment[self.nominal_value_idx].value

    def _get_fair_value(self, investment: list, multiplier=1000) -> float:
        return investment[self.fair_value_idx].value * multiplier

    def _get_currency(self, investment: list) -> str:
        return investment[self.currency_col].value

    def _get_company_name(self, investment: list) -> str:
        return investment[self.company_name_idx].value

    def _get_company_category(self, investment: list) -> str:
        return investment[self.company_category_idx].value

    @property
    def _investments_rows(self) -> Generator[list, None, None]:
        if self._file_ext == ".xls":
            return self._get_rows_from_xls(self._sheet, self.COMPANIES_START_ROW_IDX, self._sheet.nrows,  # type: ignore
                                           self._get_company_id)
        if self._file_ext == ".xlsx":
            return self._get_rows_from_xlsx(self._sheet, self.COMPANIES_START_ROW_IDX, self._sheet.max_row,
                                            # type: ignore
                                            self._get_company_id)
        raise ValueError(f"No defined way to extract rows from {self._file_ext} file")

    @property
    def investments(self) -> Generator[CompanyInvestment, None, None]:
        return (CompanyInvestment(issuer_id=self._get_company_id(investment),
                                  nominal_value=self._get_nominal_value(investment),
                                  currency=self._get_currency(investment),
                                  name=self._get_company_name(investment),
                                  category=self._get_company_category(investment)) for investment in
                self._investments_rows)

    @property
    def summed_investments(self) -> InvestmentPortfolio:
        companies_dict = {}  # type: Dict[Tuple[str, str], CompanyInvestment]
        for investment in self.investments:
            companies_dict[(investment.issuer_id, investment.currency)] = companies_dict.get(
                (investment.issuer_id, investment.currency),
                CompanyInvestment(
                    issuer_id=investment.issuer_id,
                    nominal_value=0,
                    currency=investment.currency,
                    name=investment.name,
                    category=investment.category,
                )
            ) + investment
        return companies_dict
